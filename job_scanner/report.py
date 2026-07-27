"""Build the Excel report and email it."""
from __future__ import annotations

import datetime as dt
import os
import smtplib
import ssl
from email.message import EmailMessage

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .sources import Job

HEADERS = ["#", "Match %", "Title", "Company", "Location", "Exp needed",
           "Source", "Posted", "Recommended resume", "Why it fits (keywords)",
           "Apply link"]


def build_excel(jobs: list[Job], path: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Your List"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    for col, name in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=col, value=name)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(vertical="center", wrap_text=True)

    for i, j in enumerate(jobs, start=1):
        row = i + 1
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=j.score)
        ws.cell(row=row, column=3, value=j.title)
        ws.cell(row=row, column=4, value=j.company)
        ws.cell(row=row, column=5, value=j.location)
        ws.cell(row=row, column=6, value=j.experience_req)
        ws.cell(row=row, column=7, value=j.source)
        ws.cell(row=row, column=8, value=j.posted)
        ws.cell(row=row, column=9, value=j.resume)
        ws.cell(row=row, column=10, value=", ".join(j.matched[:8]))
        link = ws.cell(row=row, column=11, value=j.url or "")
        if j.url:
            link.hyperlink = j.url
            link.font = Font(color="0563C1", underline="single")
        # colour the score cell green->amber
        fill = "C6EFCE" if j.score >= 60 else ("FFEB9C" if j.score >= 40 else "FCE4D6")
        ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor=fill)

    widths = [4, 8, 38, 24, 20, 11, 9, 15, 34, 30, 44]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(jobs) + 1}"

    wb.save(path)
    return path


def send_email(path: str, jobs: list[Job], subject: str, to_addr: str,
               from_addr: str) -> None:
    """Send via Gmail SMTP. Needs GMAIL_APP_PASSWORD in the environment."""
    app_pw = os.environ.get("GMAIL_APP_PASSWORD")
    gmail = os.environ.get("GMAIL_ADDRESS") or from_addr
    from_addr = from_addr or gmail
    to_addr = to_addr or gmail
    if not (app_pw and gmail):
        raise RuntimeError(
            "Set GMAIL_ADDRESS and GMAIL_APP_PASSWORD env vars to send email."
        )

    today = dt.date.today().strftime("%d %b %Y")
    top = jobs[:5]
    lines = "\n".join(
        f"  {j.score}%  {j.title} — {j.company}\n"
        f"        exp: {j.experience_req} | apply with: {j.resume}" for j in top
    ) or "  (no new matches in the last 24h)"
    body = (
        f"Hi Ajinkya,\n\n"
        f"Here are {len(jobs)} India-based marketing & strategy roles (0-3 yrs "
        f"experience) posted in the last 24 hours that fit your resume, best "
        f"matches first. The full list — with the resume to apply with and the "
        f"apply link for each — is in the attached Excel.\n\n"
        f"Top picks:\n{lines}\n\n"
        f"Generated automatically on {today}.\n"
    )

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = from_addr
    msg["To"] = to_addr
    msg.set_content(body)

    with open(path, "rb") as f:
        msg.add_attachment(
            f.read(),
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=os.path.basename(path),
        )

    ctx = ssl.create_default_context()
    with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=ctx) as s:
        s.login(gmail, app_pw)
        s.send_message(msg)
